import numpy as np
import random
from datetime import datetime, timedelta
import faker
import csv
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Initialize faker
fake = faker.Faker()
random.seed(42)  # For reproducibility

# Constants from the original data
LOAN_TYPES = ["SNAP", "DALEX", "SWIFT"]
LOAN_TERMS = [12, 24, 30, 36, 42, 48, 54, 60, 78]
LOCATION_CODES = ["NS100", "CS100", "WA", "SS101", "CAPE COAST", "KOFORIDUA", "HO", "TAMALE", "KUMASI", "ACCRA", "TAKORADI", "SUNYANI"]
DEDUCTION_STATUS = ["Full Payment", "Unexpected Payment", "Not Due", "Under Payment", "Over Payment", "Non Payment"]
TITLES = ["Mr.", "Mrs.", "Dr.", "Prof.", "Miss"]
MARITAL_STATUS = ["Single", "Married", "Divorced", "Widowed"]
GENDERS = ["Male", "Female", "Other"]

# Generate a consistent set of employee IDs for both datasets
def generate_employee_ids(count):
    employee_ids = []
    for _ in range(count):
        num_digits = random.choice([5, 6, 7])
        numeric_part = ''.join(random.choices('0123456789', k=num_digits))
        employee_ids.append(f"{numeric_part}CR")
    return employee_ids

# Generate some additional IDs for the issues we want to create
def generate_additional_ids(count, existing_ids):
    new_ids = []
    for _ in range(count):
        num_digits = random.choice([5, 6, 7])
        numeric_part = ''.join(random.choices('0123456789', k=num_digits))
        new_id = f"{numeric_part}CR"
        # Make sure we don't accidentally generate an existing ID
        while new_id in existing_ids:
            numeric_part = ''.join(random.choices('0123456789', k=num_digits))
            new_id = f"{numeric_part}CR"
        new_ids.append(new_id)
    return new_ids

# Random date between two dates
def random_date(start, end):
    delta = end - start
    int_delta = (delta.days * 24 * 60 * 60) + delta.seconds
    random_second = random.randrange(int_delta)
    return start + timedelta(seconds=random_second)

# Format date as MM/DD/YYYY
def format_date(date):
    return date.strftime("%m/%d/%Y")

# Format period as MMMYYYY
def format_period(date):
    months = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
    return f"{months[date.month - 1]}{date.year}"

# Generate Dalex Paddy ID
def generate_dalex_paddy():
    region = random.choice(["N", "S"])
    section1 = f"{random.randint(0, 99):02d}"
    section2 = f"{random.randint(0, 999):03d}"
    section3 = f"{random.randint(0, 99999):05d}"
    return f"{region}-{section1}-{section2}-{section3}"

# Generate Team Leader ID
def generate_team_leader(dalex_paddy):
    if not dalex_paddy:
        return ""
    parts = dalex_paddy.split('-')
    return f"{parts[0]}-{parts[1]}-{parts[2]}-00000"

# Calculate financial fields based on loan amount and term
def calculate_financial_fields(loan_amount, loan_term, loan_type):
    # Administrative Fees (10-12% of loan amount)
    if loan_type == "SNAP":
        admin_fee_percent = random.choice([10, 12])
    elif loan_type == "DALEX":
        admin_fee_percent = random.choice([6, 12])
    else:  # SWIFT
        admin_fee_percent = 12
    
    admin_fees = round(loan_amount * admin_fee_percent / 100)
    
    # Total Interest (varies by term and loan type)
    # Using a simple model: Interest = Loan Amount * (Term / 12) * interest rate
    if loan_type == "SNAP":
        interest_rate_yearly = random.uniform(0.2, 0.25)
    elif loan_type == "DALEX":
        interest_rate_yearly = random.uniform(0.18, 0.24)
    else:  # SWIFT
        interest_rate_yearly = random.uniform(0.19, 0.22)
    
    total_interest = loan_amount * (loan_term / 12) * interest_rate_yearly
    rounded_total_interest = round(total_interest, 10)
    
    # Total Collectible
    total_collectible = loan_amount + admin_fees + rounded_total_interest
    
    # Net Loan Amount (sometimes 0, sometimes matches loan amount)
    net_loan_amount = random.choice([
        0, 
        loan_amount, 
        round(loan_amount * random.uniform(0.2, 0.9), 2)
    ])
    
    # Monthly Installment
    monthly_installment = round(total_collectible / loan_term, 2)
    
    return {
        'admin_fees': admin_fees,
        'total_interest': rounded_total_interest,
        'total_collectible': total_collectible,
        'net_loan_amount': net_loan_amount,
        'monthly_installment': monthly_installment
    }

# Generate loan progress data
def generate_loan_progress(loan_amount, total_interest, total_collectible, loan_issue_date, current_date, monthly_installment):
    # Calculate months elapsed since loan issue
    loan_start_date = loan_issue_date
    months_elapsed = (current_date.year - loan_start_date.year) * 12 + (current_date.month - loan_start_date.month)
    
    # Decide payment progress based on months elapsed and some randomness
    progress_percent = min(1, months_elapsed / 48) * random.uniform(0.8, 1.2)
    progress_percent = min(progress_percent, 1)  # Cap at 100%
    
    principal_due = loan_amount * progress_percent
    interest_due = total_interest * progress_percent
    total_due = principal_due + interest_due
    
    # Payment consistency
    payment_consistency = random.choice(["full", "partial", "over"])
    
    if payment_consistency == "full":
        principal_paid = principal_due
        interest_paid = interest_due
    elif payment_consistency == "partial":
        payment_factor = random.uniform(0.8, 0.99)
        principal_paid = principal_due * payment_factor
        interest_paid = interest_due * payment_factor
    else:  # over
        payment_factor = random.uniform(1.01, 1.1)
        principal_paid = min(principal_due * payment_factor, loan_amount)
        interest_paid = min(interest_due * payment_factor, total_interest)
    
    total_paid = principal_paid + interest_paid
    
    # Outstanding loan balance
    outstanding_balance = total_collectible - total_paid
    
    # Accumulated arrears (only if partial payment)
    accumulated_arrears = total_due - total_paid if payment_consistency == "partial" else 0
    
    # NDIA (Non-Deduction In Arrears) - random for partial payments
    ndia = random.choice([0, random.uniform(0, 500)]) if payment_consistency == "partial" else 0
    
    # Deduction status based on payment consistency
    if payment_consistency == "full":
        deduction_status = "Full Payment"
    elif payment_consistency == "partial":
        deduction_status = random.choice(["Under Payment", "Non Payment"])
    else:  # over
        deduction_status = "Over Payment"
    
    # If loan is not due yet
    if months_elapsed <= 0:
        return {
            'principal_due': 0,
            'interest_due': 0,
            'total_due': 0,
            'principal_paid': 0,
            'interest_paid': 0,
            'total_paid': 0,
            'principal_paid2': 0,
            'interest_paid2': 0,
            'total_paid2': 0,
            'paid': False,
            'cancelled': False,
            'outstanding_loan_balance': total_collectible,
            'accumulated_arrears': 0,
            'ndia': 0,
            'prevailing_posted_repayment': 0,
            'prevailing_due_payment': 0,
            'current_missed_deduction': 0,
            'admin_charge': 0,
            'recovery_rate': 0,
            'deduction_status': "Not Due"
        }
    
    # Admin charge and recovery rate
    admin_charge = round(loan_amount * 0.001, 4)
    recovery_rate = random.uniform(20, 90) if payment_consistency == "partial" else 100
    
    return {
        'principal_due': round(principal_due, 2),
        'interest_due': round(interest_due, 2),
        'total_due': round(total_due, 2),
        'principal_paid': round(principal_paid, 2),
        'interest_paid': round(interest_paid, 2),
        'total_paid': round(total_paid, 2),
        'principal_paid2': 0,
        'interest_paid2': -round(random.uniform(100, 2500), 2) if deduction_status == "Over Payment" else 0,
        'total_paid2': -round(random.uniform(100, 2500), 2) if deduction_status == "Over Payment" else 0,
        'paid': False,
        'cancelled': False,
        'outstanding_loan_balance': outstanding_balance,
        'accumulated_arrears': accumulated_arrears if accumulated_arrears > 0 else 0,
        'ndia': ndia if ndia > 0 else 0,
        'prevailing_posted_repayment': monthly_installment,
        'prevailing_due_payment': monthly_installment,
        'current_missed_deduction': (
            monthly_installment if deduction_status == "Non Payment" else
            round(monthly_installment * random.uniform(0.1, 0.5), 2) if deduction_status == "Under Payment" else
            round(random.uniform(-0.01, 0.01), 2)
        ),
        'admin_charge': admin_charge,
        'recovery_rate': (
            round(100 + random.uniform(0.0001, 0.01), 10) if deduction_status == "Over Payment" else recovery_rate
        ),
        'deduction_status': deduction_status
    }

# Generate client data with duplications and missing data
def generate_client_data(employee_ids, count, extra_ids=None):
    clients = []
    
    # Date ranges
    start_dob = datetime(1965, 1, 1)
    end_dob = datetime(2003, 12, 31)
    start_employment = datetime(2015, 1, 1)
    end_employment = datetime(2023, 12, 31)
    
    # Potential employers
    employers = [
        "Green-Moore", "Ball-Harrison", "Wagner-Mitchell", 
        "Shepard, Knight and Simon", "Lopez, Richardson and Wolf",
        "Collins-Taylor", "Lewis, Benitez and Ferguson",
        "Ashley-Harrison", "Scott-Martin", "Mcclure-Frost",
        "Shepherd-Cook", "Reed-Flores", "Davis Ltd",
        "Huynh-Torres", "Cooper, Stevens and Dunn",
        "Bonilla Inc", "Thomas-James", "Holmes, Dunn and Martinez",
        "Wang Inc", "Johnson-Smith", "Peterson Group",
        "Lee, Adams and White", "Hall-Miller"
    ]
    
    # Create a set of addresses to reuse for duplicates (issue #1)
    duplicate_addresses = [fake.address() for _ in range(int(count * 0.05))]  # 5% of records will have duplicate addresses
    duplicate_dobs = [random_date(start_dob, end_dob) for _ in range(int(count * 0.05))]  # 5% will have duplicate DOBs
    
    # Process the regular employee IDs
    for i in range(count):
        if i % 1000 == 0:
            print(f"Generating client {i+1}/{count}")
            
        employee_id = employee_ids[i]
        last_name = fake.last_name()
        other_names = fake.first_name()
        
        gender = random.choice(GENDERS)
        title = random.choice(TITLES)
        marital_status = random.choice(MARITAL_STATUS)
        
        # Issue #5: Missing date of birth for some records (about 3%)
        if random.random() < 0.03:
            dob = ""
        else:
            # Issue #1: Duplicate DOB for some records
            if random.random() < 0.05:
                dob = random.choice(duplicate_dobs).strftime("%Y-%m-%d")
            else:
                dob = random_date(start_dob, end_dob).strftime("%Y-%m-%d")
        
        employment_date = random_date(start_employment, end_employment).strftime("%Y-%m-%d")
        
        # Issue #1: Duplicate addresses for some records
        if random.random() < 0.05:
            residential_address = random.choice(duplicate_addresses)
        else:
            residential_address = fake.address()
        
        postal_address = fake.address()
        phone_no = fake.phone_number()
        
        employer = random.choice(employers)
        prev_employee_no = ''.join(random.choices('0123456789', k=4))
        
        # SSN in format XXX-XX-XXXX
        ssn1 = ''.join(random.choices('0123456789', k=3))
        ssn2 = ''.join(random.choices('0123456789', k=2))
        ssn3 = ''.join(random.choices('0123456789', k=4))
        social_security_no = f"{ssn1}-{ssn2}-{ssn3}"
        
        # Voter ID in format VOTXXXXXX
        voter_id_no = f"VOT{''.join(random.choices('0123456789', k=6))}"
        
        # Next of kin details
        next_of_kin = fake.name()
        next_of_kin_contact = fake.phone_number()
        next_of_kin_address = fake.address()
        
        # Search name
        search_name = f"{last_name} {other_names}"
        
        clients.append({
            "Employee Id": employee_id,
            "Lastname": last_name,
            "Othernames": other_names,
            "Residential Address": residential_address,
            "Postal Address": postal_address,
            "Client Phone No.": phone_no,
            "Title": title,
            "Marital Status": marital_status,
            "Gender": gender,
            "Date of Birth": dob,
            "Employer": employer,
            "Previous Employee No.": prev_employee_no,
            "Social Security No.": social_security_no,
            "Voters Id No.": voter_id_no,
            "Employment Date": employment_date,
            "Next of Kin": next_of_kin,
            "Next of Kin Contact:": next_of_kin_contact,
            "Next of Kin Address": next_of_kin_address,
            "Search Name": search_name
        })
    
    # Process the extra IDs for issue #3: clients without corresponding loans
    if extra_ids:
        print(f"Adding {len(extra_ids)} extra clients without matching loans...")
        for employee_id in extra_ids:
            last_name = fake.last_name()
            other_names = fake.first_name()
            
            gender = random.choice(GENDERS)
            title = random.choice(TITLES)
            marital_status = random.choice(MARITAL_STATUS)
            
            # No missing DOB for these records
            dob = random_date(start_dob, end_dob).strftime("%Y-%m-%d")
            employment_date = random_date(start_employment, end_employment).strftime("%Y-%m-%d")
            
            residential_address = fake.address()
            postal_address = fake.address()
            phone_no = fake.phone_number()
            
            employer = random.choice(employers)
            prev_employee_no = ''.join(random.choices('0123456789', k=4))
            
            # SSN in format XXX-XX-XXXX
            ssn1 = ''.join(random.choices('0123456789', k=3))
            ssn2 = ''.join(random.choices('0123456789', k=2))
            ssn3 = ''.join(random.choices('0123456789', k=4))
            social_security_no = f"{ssn1}-{ssn2}-{ssn3}"
            
            # Voter ID in format VOTXXXXXX
            voter_id_no = f"VOT{''.join(random.choices('0123456789', k=6))}"
            
            # Next of kin details
            next_of_kin = fake.name()
            next_of_kin_contact = fake.phone_number()
            next_of_kin_address = fake.address()
            
            # Search name
            search_name = f"{last_name} {other_names}"
            
            clients.append({
                "Employee Id": employee_id,
                "Lastname": last_name,
                "Othernames": other_names,
                "Residential Address": residential_address,
                "Postal Address": postal_address,
                "Client Phone No.": phone_no,
                "Title": title,
                "Marital Status": marital_status,
                "Gender": gender,
                "Date of Birth": dob,
                "Employer": employer,
                "Previous Employee No.": prev_employee_no,
                "Social Security No.": social_security_no,
                "Voters Id No.": voter_id_no,
                "Employment Date": employment_date,
                "Next of Kin": next_of_kin,
                "Next of Kin Contact:": next_of_kin_contact,
                "Next of Kin Address": next_of_kin_address,
                "Search Name": search_name
            })
    
    return clients

# Generate loan data with intentional issues
def generate_loan_data(employee_ids, client_data, count, extra_loan_ids=None):
    loans = []
    current_date = datetime(2025, 4, 7)  # April 7, 2025
    start_loan_date = datetime(2016, 1, 1)
    end_loan_date = datetime(2025, 3, 31)  # End of March 2025
    
    # Start loan numbers from L100000
    loan_counter = 100000
    
    # Create a set to track used loan numbers (for duplicate check)
    used_loan_numbers = set()
    
    # Set of loan numbers to duplicate (issue #2) - about 3% of loans
    duplicate_loan_numbers = set()
    
    for i in range(count):
        if i % 1000 == 0:
            print(f"Generating loan {i+1}/{count}")
            
        employee_id = employee_ids[i]
        client = next((c for c in client_data if c["Employee Id"] == employee_id), None)
        
        # Issue #4: Some loans don't match to valid client data (about 3%)
        if client and random.random() > 0.03:
            last_name = client["Lastname"]
            other_names = client["Othernames"]
            employee_name = f"{last_name.upper()} {other_names.upper()}"
        else:
            employee_name = f"{fake.last_name().upper()} {fake.first_name().upper()}"
        
        # Generate a loan number
        loan_no = f"L{loan_counter}"
        loan_counter += 1
        
        # Issue #2: Duplicate some loan numbers (about 3%)
        if random.random() < 0.03:
            duplicate_loan_numbers.add(loan_no)
        
        used_loan_numbers.add(loan_no)
        
        # Loan dates
        loan_issue_date = random_date(start_loan_date, end_loan_date)
        
        # Loan details
        loan_type = random.choice(LOAN_TYPES)
        loan_term = random.choice(LOAN_TERMS)
        loan_amount = random.randint(1000, 20000)
        
        # Calculate subsequent dates
        deduction_start_date = (loan_issue_date.replace(day=1) + timedelta(days=32)).replace(day=1)
        submission_period_date = loan_issue_date.replace(day=1)
        maturity_period_date = deduction_start_date + timedelta(days=30*loan_term)
        
        # Format date strings
        loan_issue_date_str = format_date(loan_issue_date)
        deduction_start_period = format_period(deduction_start_date)
        submission_period = format_period(submission_period_date)
        maturity_period = format_period(maturity_period_date)
        
        # Location and team details
        location_code = random.choice(LOCATION_CODES)
        dalex_paddy = generate_dalex_paddy() if random.random() > 0.1 else ""
        team_leader = generate_team_leader(dalex_paddy)
        
        # Financial calculations
        financial = calculate_financial_fields(loan_amount, loan_term, loan_type)
        
        # Loan progress data
        progress = generate_loan_progress(
            loan_amount, 
            financial['total_interest'], 
            financial['total_collectible'], 
            loan_issue_date, 
            current_date,
            financial['monthly_installment']
        )
        
        loans.append({
            "Loan No.": loan_no,
            "Employee Id": employee_id,
            "Employee Name": employee_name,
            "Employer": "CAGD",
            "Loan Issue Date": loan_issue_date_str,
            "Deduction Start Period": deduction_start_period,
            "Submission Period": submission_period,
            "Maturity Period": maturity_period,
            "Location Code": location_code,
            "Dalex Paddy": dalex_paddy,
            "Team Leader": team_leader,
            "Loan Type": loan_type,
            "Loan Amount": f"{loan_amount:.2f}",
            "Loan Term": str(loan_term),
            "Administrative Fees": f"{financial['admin_fees']:.2f}",
            "Total Interest": f"{financial['total_interest']:.2f}",
            "Total Collectible": f"{financial['total_collectible']:.2f}",
            "Net Loan Amount": f"{financial['net_loan_amount']:.2f}",
            "Monthly Installment": f"{financial['monthly_installment']:.2f}",
            "Principal Due": str(progress['principal_due']),
            "Interest Due": str(progress['interest_due']),
            "Total Due": str(progress['total_due']),
            "Principal Paid": str(progress['principal_paid']),
            "Interest Paid": str(progress['interest_paid']),
            "Total Paid": str(progress['total_paid']),
            "Principal Paid2": str(progress['principal_paid2']),
            "Interest Paid2": str(progress['interest_paid2']),
            "Total Paid2": str(progress['total_paid2']),
            "Paid": progress['paid'],
            "Cancelled": progress['cancelled'],
            "Outstanding Loan Balance": progress['outstanding_loan_balance'],
            "Accumulated Arrears": progress['accumulated_arrears'],
            "NDIA": progress['ndia'],
            "Prevailing Posted Repayment": str(progress['prevailing_posted_repayment']),
            "Prevailing Due Payment": str(progress['prevailing_due_payment']),
            "Current Missed Deduction": str(progress['current_missed_deduction']),
            "Admin Charge": str(progress['admin_charge']),
            "Recovery Rate": str(progress['recovery_rate']),
            "Deduction Status": progress['deduction_status']
        })
    
    # Process the extra employee IDs (Issue #4: loans without matching client data)
    if extra_loan_ids:
        print(f"Adding {len(extra_loan_ids)} extra loans without matching clients...")
        for employee_id in extra_loan_ids:
            employee_name = f"{fake.last_name().upper()} {fake.first_name().upper()}"
            
            # Generate a loan number
            loan_no = f"L{loan_counter}"
            loan_counter += 1
            used_loan_numbers.add(loan_no)
            
            # Loan dates
            loan_issue_date = random_date(start_loan_date, end_loan_date)
            
            # Loan details
            loan_type = random.choice(LOAN_TYPES)
            loan_term = random.choice(LOAN_TERMS)
            loan_amount = random.randint(1000, 20000)
            
            # Calculate subsequent dates
            deduction_start_date = (loan_issue_date.replace(day=1) + timedelta(days=32)).replace(day=1)
            submission_period_date = loan_issue_date.replace(day=1)
            maturity_period_date = deduction_start_date + timedelta(days=30*loan_term)
            
            # Format date strings
            loan_issue_date_str = format_date(loan_issue_date)
            deduction_start_period = format_period(deduction_start_date)
            submission_period = format_period(submission_period_date)
            maturity_period = format_period(maturity_period_date)
            
            # Location and team details
            location_code = random.choice(LOCATION_CODES)
            dalex_paddy = generate_dalex_paddy() if random.random() > 0.1 else ""
            team_leader = generate_team_leader(dalex_paddy)
            
            # Financial calculations
            financial = calculate_financial_fields(loan_amount, loan_term, loan_type)
            
            # Loan progress data
            progress = generate_loan_progress(
                loan_amount, 
                financial['total_interest'], 
                financial['total_collectible'], 
                loan_issue_date, 
                current_date,
                financial['monthly_installment']
            )
            
            loans.append({
                "Loan No.": loan_no,
                "Employee Id": employee_id,
                "Employee Name": employee_name,
                "Employer": "CAGD",
                "Loan Issue Date": loan_issue_date_str,
                "Deduction Start Period": deduction_start_period,
                "Submission Period": submission_period,
                "Maturity Period": maturity_period,
                "Location Code": location_code,
                "Dalex Paddy": dalex_paddy,
                "Team Leader": team_leader,
                "Loan Type": loan_type,
                "Loan Amount": f"{loan_amount:.2f}",
                "Loan Term": str(loan_term),
                "Administrative Fees": f"{financial['admin_fees']:.2f}",
                "Total Interest": f"{financial['total_interest']:.2f}",
                "Total Collectible": f"{financial['total_collectible']:.2f}",
                "Net Loan Amount": f"{financial['net_loan_amount']:.2f}",
                "Monthly Installment": f"{financial['monthly_installment']:.2f}",
                "Principal Due": str(progress['principal_due']),
                "Interest Due": str(progress['interest_due']),
                "Total Due": str(progress['total_due']),
                "Principal Paid": str(progress['principal_paid']),
                "Interest Paid": str(progress['interest_paid']),
                "Total Paid": str(progress['total_paid']),
                "Principal Paid2": str(progress['principal_paid2']),
                "Interest Paid2": str(progress['interest_paid2']),
                "Total Paid2": str(progress['total_paid2']),
                "Paid": progress['paid'],
                "Cancelled": progress['cancelled'],
                "Outstanding Loan Balance": progress['outstanding_loan_balance'],
                "Accumulated Arrears": progress['accumulated_arrears'],
                "NDIA": progress['ndia'],
                "Prevailing Posted Repayment": str(progress['prevailing_posted_repayment']),
                "Prevailing Due Payment": str(progress['prevailing_due_payment']),
                "Current Missed Deduction": str(progress['current_missed_deduction']),
                "Admin Charge": str(progress['admin_charge']),
                "Recovery Rate": str(progress['recovery_rate']),
                "Deduction Status": progress['deduction_status']
            })
    
    # Issue #2: Add duplicate loan numbers 
    print(f"Adding {len(duplicate_loan_numbers)} loan records with duplicate IDs...")
    for loan_no in duplicate_loan_numbers:
        # Find the original loan with this number
        original_loan = next(loan for loan in loans if loan["Loan No."] == loan_no)
        # Create a copy with some variations
        duplicate_loan = original_loan.copy()
        duplicate_loan["Loan Issue Date"] = format_date(random_date(start_loan_date, end_loan_date))
        
        # Parse the original loan amount, add a random value, and format it back
        original_amount = float(original_loan["Loan Amount"].replace(',', ''))
        new_amount = original_amount + random.randint(-1000, 1000)
        duplicate_loan["Loan Amount"] = f"{new_amount:.2f}"
        
        loans.append(duplicate_loan)
    
    return loans

# Function to write data to CSV files in batches
def write_to_csv(data, file_path, batch_size=10000):
    print(f"Writing data to CSV file {file_path}...")
    
    # Write header first
    with open(file_path, 'w', newline='') as f:
        writer = csv.writer(f, delimiter='\t')
        writer.writerow(data[0].keys())
    
    # Then append data in batches
    total_records = len(data)
    for start_idx in range(0, total_records, batch_size):
        end_idx = min(start_idx + batch_size, total_records)
        batch = data[start_idx:end_idx]
        
        with open(file_path, 'a', newline='') as f:
            writer = csv.writer(f, delimiter='\t')
            for record in batch:
                writer.writerow(record.values())
        
        print(f"Wrote records {start_idx + 1} to {end_idx} to CSV file")
    
    print(f"Successfully wrote {total_records} records to {file_path}")

# Function to write data to XLSX files in batches - optimized for large datasets
def write_to_excel(data, file_path, sheet_name, batch_size=10000):
    print(f"Writing data to Excel file {file_path}...")
    
    # Convert the first batch to DataFrame to get column information
    first_batch = data[:min(batch_size, len(data))]
    df_first = pd.DataFrame(first_batch)
    
    # Create a new Excel file with headers only
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df_first.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"Created Excel file with headers at {file_path}")
    
    # Now append data in batches, skipping the first batch which is already written
    total_records = len(data)
    
    if total_records <= batch_size:
        print(f"All {total_records} records written in a single batch")
        return
    
    for start_idx in range(batch_size, total_records, batch_size):
        end_idx = min(start_idx + batch_size, total_records)
        batch = data[start_idx:end_idx]
        
        # Convert batch to DataFrame
        df_batch = pd.DataFrame(batch)
        
        # Open the workbook
        book = load_workbook(file_path)
        sheet = book[sheet_name]
        
        # Write data rows
        for r_idx, row in enumerate(dataframe_to_rows(df_batch, index=False, header=False)):
            for c_idx, value in enumerate(row):
                sheet.cell(row=start_idx+r_idx+2, column=c_idx+1, value=value)
        
        # Save workbook
        book.save(file_path)
        print(f"Wrote records {start_idx + 1} to {end_idx} to Excel file")
            
    print(f"Successfully wrote {total_records} records to {file_path}")

# Main function to generate the data
def generate_data(client_count=70000, loan_count=70000, batch_size=10000, output_dir=".", output_format="xlsx"):
    """
    Generate dummy loan and client data with intentional issues.
    
    Parameters:
    client_count (int): Number of client records to generate
    loan_count (int): Number of loan records to generate
    batch_size (int): Batch size for writing data to files
    output_dir (str): Directory to save output files
    output_format (str): Format to save data - 'csv', 'xlsx', or 'both'
    
    Returns:
    dict: Statistics about the generated data
    """
    # Check if output format is valid
    if output_format not in ['csv', 'xlsx', 'both']:
        raise ValueError("output_format must be one of: 'csv', 'xlsx', 'both'")
    
    # First, generate the consistent set of employee IDs
    base_employee_ids = generate_employee_ids(client_count)
    
    # Generate about 5% extra IDs for each dataset to create the mismatches
    client_extra_ids = generate_additional_ids(int(client_count * 0.05), base_employee_ids)
    loan_extra_ids = generate_additional_ids(int(loan_count * 0.05), base_employee_ids + client_extra_ids)
    
    # Generate client data
    print(f"Generating {client_count} client records...")
    client_data = generate_client_data(base_employee_ids, client_count, client_extra_ids)
    
    # Generate loan data 
    print(f"Generating {loan_count} loan records...")
    loan_data = generate_loan_data(base_employee_ids, client_data, loan_count, loan_extra_ids)
    
    # Set file paths
    csv_client_path = os.path.join(output_dir, 'client_data_70k.csv')
    csv_loan_path = os.path.join(output_dir, 'loan_data_70k.csv')
    excel_client_path = os.path.join(output_dir, 'client_data_70k.xlsx')
    excel_loan_path = os.path.join(output_dir, 'loan_data_70k.xlsx')
    
    # Write data to files based on the specified format
    if output_format in ['csv', 'both']:
        write_to_csv(client_data, csv_client_path, batch_size)
        write_to_csv(loan_data, csv_loan_path, batch_size)
        print(f"CSV data saved to:\n- {csv_client_path}\n- {csv_loan_path}")
        
    if output_format in ['xlsx', 'both']:
        write_to_excel(client_data, excel_client_path, 'Clients', batch_size)
        write_to_excel(loan_data, excel_loan_path, 'Loans', batch_size)
        print(f"Excel data saved to:\n- {excel_client_path}\n- {excel_loan_path}")
    
    print("Data generation complete!")
    
    # Print summary of issues created
    duplicate_loan_ids = {x: sum(1 for loan in loan_data if loan['Loan No.'] == x) for x in set(loan['Loan No.'] for loan in loan_data)}
    duplicate_loan_ids = {k: v for k, v in duplicate_loan_ids.items() if v > 1}
    
    print("\nIssues summary:")
    print(f"1. Clients with duplicate addresses/DOB: ~{int(client_count * 0.05)} records")
    print(f"2. Duplicate loan IDs: {len(duplicate_loan_ids)} unique IDs duplicated")
    print(f"3. Clients without matching loans: ~{len(client_extra_ids)} records")
    print(f"4. Loans without matching clients: ~{len(loan_extra_ids)} records")
    print(f"5. Clients with missing DOB: ~{int(client_count * 0.03)} records")
    
    # Return diagnostics
    return {
        'client_count': len(client_data),
        'loan_count': len(loan_data),
        'duplicate_loan_ids': len(duplicate_loan_ids),
        'clients_without_loans': len(client_extra_ids),
        'loans_without_clients': len(loan_extra_ids)
    }

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate dummy loan and client data with intentional issues.')
    parser.add_argument('--client-count', type=int, default=70000, 
                        help='Number of client records to generate (default: 70000)')
    parser.add_argument('--loan-count', type=int, default=70000, 
                        help='Number of loan records to generate (default: 70000)')
    parser.add_argument('--batch-size', type=int, default=10000, 
                        help='Batch size for writing data to manage memory (default: 10000)')
    parser.add_argument('--output-dir', type=str, default='.', 
                        help='Directory to save output files (default: current directory)')
    parser.add_argument('--format', type=str, choices=['csv', 'xlsx', 'both'], default='xlsx',
                        help='Output format: csv, xlsx, or both (default: xlsx)')
    
    args = parser.parse_args()
    
    # Create output directory if it doesn't exist
    if not os.path.exists(args.output_dir):
        os.makedirs(args.output_dir)
    
    # Run data generation
    result = generate_data(args.client_count, args.loan_count, args.batch_size, args.output_dir, args.format)
    
    print("\nGeneration complete with the following statistics:")
    for key, value in result.items():
        print(f"{key}: {value}")
