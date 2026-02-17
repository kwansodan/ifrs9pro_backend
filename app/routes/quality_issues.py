from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    status,
    Body,
    Query,
)
from openpyxl import Workbook
from sqlalchemy.orm import joinedload
from io import BytesIO
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Dict, Any, Optional
from fastapi import HTTPException, Depends, status
from datetime import datetime
from io import BytesIO
import pandas as pd
import logging

from app.database import get_db
from app.dependencies import get_tenant_db
from app.models import Portfolio, User, QualityIssue, QualityIssueComment
from app.auth.utils import get_current_active_user
from app.schemas import (
    QualityIssueResponse,
    QualityIssueUpdate,
    QualityIssueCommentCreate,
    QualityIssueCommentModel,
    QualityCheckSummary,
    QualityIssueSummary,
)
from app.utils.quality_checks import create_quality_issues_if_needed

# Create a separate router for quality issues
router = APIRouter(prefix="/portfolios", tags=["quality-issues"])


def transform_affected_records(quality_issues: List[QualityIssue]) -> List[QualityIssue]:
    """
    Transform affected_records from dictionary to list format for backward compatibility.
    This is needed because the schema expects affected_records to be a list of dictionaries,
    but older records in the database might have it as a single dictionary.
    """
    logger = logging.getLogger(__name__)
    
    for issue in quality_issues:
        try:
            # Check if affected_records is a dictionary (old format)
            if isinstance(issue.affected_records, dict):
                # Convert it to a list containing that dictionary
                issue.affected_records = [issue.affected_records]
                logger.debug(f"Transformed affected_records for issue {issue.id} to list format")
        except Exception as e:
            logger.error(f"Error transforming affected_records for issue {issue.id}: {str(e)}")
    
    return quality_issues


def excel_safe(values):
    """
    Convert SQLAlchemy rows to list/tuple and safely handle complex values (dict/list)
    for OpenPyXL write-only mode.
    """
    if not isinstance(values, (list, tuple)):
        values = tuple(values)
        
    return [
        str(v) if isinstance(v, (dict, list)) else v
        for v in values
    ]


@router.get(
    "/{portfolio_id}/quality-issues",
    description="Get aggregated quality issues for a specific portfolio",
    response_model=List[QualityIssueSummary],
)
def get_quality_issues(
    portfolio_id: int,
    status_type: Optional[str] = None,
    issue_type: Optional[str] = None,
    db: Session = Depends(get_tenant_db),
    current_user: User = Depends(get_current_active_user),
):
    # --- Verify portfolio ownership
    portfolio = (
        db.query(Portfolio)
        .filter(
            Portfolio.id == portfolio_id,
            Portfolio.user_id == current_user.id,
        )
        .first()
    )
    if not portfolio:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found")

    # --- Base filters
    base_filter = [QualityIssue.portfolio_id == portfolio_id]
    if status_type:
        base_filter.append(QualityIssue.status == status_type)
    if issue_type:
        base_filter.append(QualityIssue.issue_type == issue_type)

    # --- Issue type to generic description mapping
    ISSUE_TYPE_DESCRIPTIONS = {
        "duplicate_customer_id": "Duplicate employee ID found across multiple clients",
        "duplicate_customer_ids": "Duplicate employee ID found across multiple clients",
        "duplicate_address": "Duplicate residential address found across multiple clients",
        "duplicate_addresses": "Duplicate residential address found across multiple clients",
        "duplicate_dob": "Duplicate date of birth found across multiple clients",
        "duplicate_loan_id": "Duplicate loan number found across multiple loans",
        "duplicate_loan_ids": "Duplicate loan number found across multiple loans",
        "duplicate_phone": "Duplicate phone number found across multiple clients",
        "duplicate_phones": "Duplicate phone number found across multiple clients",
        "client_without_matching_loan": "Client has no matching loan records",
        "loan_without_matching_client": "Loan has no matching client records",
        "missing_dob": "Client record is missing date of birth",
        "missing_address": "Client record is missing residential address",
        "missing_loan_number": "Loan record is missing loan number",
        "missing_loan_date": "Loan record is missing issue date",
        "missing_loan_term": "Loan record is missing term",
        "missing_interest_rate": "Loan record is missing interest rate or fees",
        "missing_loan_amount": "Loan record is missing principal amount",
        "unmatched_employee_id": "Client has no matching loan records",
        "loan_customer_mismatch": "Loan has no matching client records",
    }

    # --- Aggregated query (Grouped by issue_type and severity)
    agg_query = (
        db.query(
            func.max(QualityIssue.id).label("issue_id"),
            QualityIssue.issue_type,
            QualityIssue.severity,
            func.count(QualityIssue.id).label("occurrence_count"),
            func.min(QualityIssue.created_at).label("first_occurrence"),
            func.max(func.coalesce(QualityIssue.updated_at, QualityIssue.created_at)).label("last_occurrence"),
        )
        .filter(*base_filter)
        .group_by(QualityIssue.issue_type, QualityIssue.severity)
    )
    agg_results = agg_query.all()

    if not agg_results:
        return []

    # --- Status distribution (Grouped by issue_type and severity)
    status_query = (
        db.query(
            QualityIssue.issue_type,
            QualityIssue.severity,
            QualityIssue.status,
            func.count(QualityIssue.id).label("count"),
        )
        .filter(*base_filter)
        .group_by(QualityIssue.issue_type, QualityIssue.severity, QualityIssue.status)
    )
    status_results = status_query.all()

    # --- Build status map
    status_map = {}
    for row in status_results:
        key = (row.issue_type, row.severity)
        if key not in status_map:
            status_map[key] = {}
        status_map[key][row.status] = row.count

    # --- Build final response as Pydantic models
    response_models = []
    for row in agg_results:
        issue_type = row.issue_type
        severity = row.severity
        key = (issue_type, severity)
        
        # Get generic description or fallback to the issue type itself
        description = ISSUE_TYPE_DESCRIPTIONS.get(issue_type, issue_type.replace("_", " ").capitalize())
        
        response_models.append(
            QualityIssueSummary(
                issue_id=row.issue_id,
                description=description,
                issue_type=issue_type,
                severity=severity,
                occurrence_count=row.occurrence_count,
                first_occurrence=row.first_occurrence,
                last_occurrence=row.last_occurrence,
                statuses=status_map.get(key, {}),
            )
        )

    # --- Sort by occurrence and severity
    response_models.sort(
        key=lambda x: (x.occurrence_count, x.severity == "high", x.severity == "medium"),
        reverse=True,
    )

    # --- Convert to JSON-compatible dicts for Swagger/testclient
    return [item.model_dump() for item in response_models]


@router.get("/{portfolio_id}/quality-issues/download", 
            description="Download all quality issues for a portfolio as Excel", 
            status_code=status.HTTP_200_OK,
            responses={404: {"description": "Portfolio not found"},
                       401: {"description": "Not authenticated"}},)
def download_all_quality_issues_excel(
    portfolio_id: int,
    status_type: Optional[str] = None,
    issue_type: Optional[str] = None,
    include_comments: bool = Query(False),
    db: Session = Depends(get_tenant_db),
    current_user: User = Depends(get_current_active_user),
):
    # ---- Validate portfolio ownership
    portfolio = (
        db.query(Portfolio)
        .filter(
            Portfolio.id == portfolio_id,
            Portfolio.user_id == current_user.id
        )
        .first()
    )
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")

    # ---- Base query
    query = db.query(QualityIssue).filter(QualityIssue.portfolio_id == portfolio_id)

    if status_type:
        query = query.filter(QualityIssue.status == status_type)
    if issue_type:
        query = query.filter(QualityIssue.issue_type == issue_type)

    query = query.order_by(QualityIssue.created_at.desc())

    # ---- Prepare Excel streaming workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Issues")

    # ---- Header row
    ws.append([
        "ID", "Issue Type", "Description",
        "Severity", "Status", "Created", "Updated"
    ])

    # ---- Stream issues in chunks
    for issue in query.yield_per(1000):
        ws.append([
            issue.id,
            issue.issue_type,
            issue.description,
            issue.severity,
            issue.status,
            issue.created_at,
            issue.updated_at,
        ])

    # ---- Comments sheet (optional)
    if include_comments:
        ws_comments = wb.create_sheet("Comments")

        ws_comments.append([
            "Issue ID", "Comment ID", "User Email",
            "Comment", "Created"
        ])

        comments_query = (
            db.query(
                QualityIssueComment.quality_issue_id,
                QualityIssueComment.id,
                User.email,
                QualityIssueComment.comment,
                QualityIssueComment.created_at
            )
            .join(User, User.id == QualityIssueComment.user_id)
            .join(QualityIssue,
                  QualityIssue.id == QualityIssueComment.quality_issue_id)
            .filter(QualityIssue.portfolio_id == portfolio_id)
        )

        for row in comments_query.yield_per(1000):
            ws_comments.append(excel_safe(row))

    # ---- Save to memory buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"quality_issues_{portfolio_id}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{portfolio_id}/quality-issues/{issue_id}", 
            description="Retrieve specific quality issues for a particular portfolio", 
            response_model=QualityIssueResponse,
            responses={404: {"description": "Portfolio not found"},
                       401: {"description": "Not authenticated"}},)
def get_quality_issue(
    portfolio_id: int,
    issue_id: int,
    db: Session = Depends(get_tenant_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Retrieve a specific quality issue by ID.
    """
    # Verify portfolio exists and belongs to current user
    portfolio = (
        db.query(Portfolio)
        .filter(Portfolio.id == portfolio_id)
        .first()
    )

    if not portfolio:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found"
        )

    # Get the quality issue
    issue = (
        db.query(QualityIssue)
        .filter(QualityIssue.id == issue_id, QualityIssue.portfolio_id == portfolio_id)
        .first()
    )

    if not issue:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Quality issue not found"
        )

    # Transform affected_records from dictionary to list format if needed
    issue = transform_affected_records([issue])[0]

    return issue


@router.put("/{portfolio_id}/quality-issues/{issue_id}", 
            description="Update quality issues including approving them", 
            response_model=QualityIssueResponse,
            responses={404: {"description": "Portfolio not found"},
                       401: {"description": "Not authenticated"}},)
def update_quality_issue(
    portfolio_id: int,
    issue_id: int,
    issue_update: QualityIssueUpdate,
    db: Session = Depends(get_tenant_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Update a quality issue, including approving it (changing status to "approved").
    """
    # Verify portfolio exists and belongs to current user
    portfolio = (
        db.query(Portfolio)
        .filter(Portfolio.id == portfolio_id)
        .first()
    )

    if not portfolio:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found"
        )

    # Get the quality issue
    issue = (
        db.query(QualityIssue)
        .filter(QualityIssue.id == issue_id, QualityIssue.portfolio_id == portfolio_id)
        .first()
    )

    if not issue:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Quality issue not found"
        )

    # Update fields if provided
    update_data = issue_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(issue, key, value)

    db.commit()
    db.refresh(issue)

    # Transform affected_records from dictionary to list format if needed
    issue = transform_affected_records([issue])[0]

    return issue


@router.post(
    "/{portfolio_id}/quality-issues/{issue_id}/comments", 
    description="Add a comment to a quality issue",
    response_model=QualityIssueCommentModel,
    responses={404: {"description": "Portfolio not found"},
               401: {"description": "Not authenticated"}},
)
def add_comment_to_quality_issue(
    portfolio_id: int,
    issue_id: int,
    comment_data: QualityIssueCommentCreate,
    db: Session = Depends(get_tenant_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Add a comment to a quality issue.
    """
    # Verify portfolio exists and belongs to current user
    portfolio = (
        db.query(Portfolio)
        .filter(Portfolio.id == portfolio_id)
        .first()
    )

    if not portfolio:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found"
        )

    # Get the quality issue
    issue = (
        db.query(QualityIssue)
        .filter(QualityIssue.id == issue_id, QualityIssue.portfolio_id == portfolio_id)
        .first()
    )

    if not issue:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Quality issue not found"
        )

    # Create new comment
    new_comment = QualityIssueComment(
        quality_issue_id=issue_id, user_id=current_user.id, comment=comment_data.comment
    )

    db.add(new_comment)
    db.commit()
    db.refresh(new_comment)

    return new_comment


@router.get(
    "/{portfolio_id}/quality-issues/{issue_id}/comments", 
    description="Get all comments for a quality issue",
    response_model=List[QualityIssueCommentModel],
    responses={404: {"description": "Portfolio not found"},
               401: {"description": "Not authenticated"}},
)
def get_quality_issue_comments(
    portfolio_id: int,
    issue_id: int,
    db: Session = Depends(get_tenant_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Get all comments for a quality issue.
    """
    # Verify portfolio exists and belongs to current user
    portfolio = (
        db.query(Portfolio)
        .filter(Portfolio.id == portfolio_id)
        .first()
    )

    if not portfolio:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found"
        )

    # Get the quality issue
    issue = (
        db.query(QualityIssue)
        .filter(QualityIssue.id == issue_id, QualityIssue.portfolio_id == portfolio_id)
        .first()
    )

    if not issue:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Quality issue not found"
        )

    # Get all comments for this issue, ordered by creation date
    comments = (
        db.query(QualityIssueComment)
        .filter(QualityIssueComment.quality_issue_id == issue_id)
        .order_by(QualityIssueComment.created_at)
        .all()
    )

    return comments


@router.put(
    "/{portfolio_id}/quality-issues/{issue_id}/comments/{comment_id}", 
    description="Edit a comment on a quality issue",
    response_model=QualityIssueCommentModel,
    responses={404: {"description": "Portfolio not found"},
               401: {"description": "Not authenticated"}},
)
def edit_quality_issue_comment(
    portfolio_id: int,
    issue_id: int,
    comment_id: int,
    comment_data: QualityIssueCommentCreate,
    db: Session = Depends(get_tenant_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Edit a comment on a quality issue.
    """
    # Verify portfolio exists and belongs to current user
    portfolio = (
        db.query(Portfolio)
        .filter(Portfolio.id == portfolio_id)
        .first()
    )
    if not portfolio:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found"
        )
    
    # Get the quality issue
    issue = (
        db.query(QualityIssue)
        .filter(QualityIssue.id == issue_id, QualityIssue.portfolio_id == portfolio_id)
        .first()
    )
    if not issue:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Quality issue not found"
        )
    
    # Get the comment and verify ownership
    comment = (
        db.query(QualityIssueComment)
        .filter(
            QualityIssueComment.id == comment_id,
            QualityIssueComment.quality_issue_id == issue_id,
            QualityIssueComment.user_id == current_user.id,
        )
        .first()
    )
    if not comment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, 
            detail="Comment not found or you don't have permission to edit it"
        )
    
    # Update comment
    comment.comment = comment_data.comment
    
    db.commit()
    db.refresh(comment)
    return comment


@router.post(
    "/{portfolio_id}/quality-issues/{issue_id}/approve", 
    description="Approve a quality issue",
    response_model=QualityIssueResponse,
    responses={404: {"description": "Portfolio not found"},
               401: {"description": "Not authenticated"}},
)
def approve_quality_issue(
    portfolio_id: int,
    issue_id: int,
    comment: Optional[str] = None,
    db: Session = Depends(get_tenant_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Approve a quality issue, changing its status to "approved".
    Optionally add a comment about the approval.
    """
    # Verify portfolio exists and belongs to current user
    portfolio = (
        db.query(Portfolio)
        .filter(Portfolio.id == portfolio_id)
        .first()
    )

    if not portfolio:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found"
        )

    # Get the quality issue
    issue = (
        db.query(QualityIssue)
        .filter(QualityIssue.id == issue_id, QualityIssue.portfolio_id == portfolio_id)
        .first()
    )

    if not issue:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Quality issue not found"
        )

    # Update status to approved
    issue.status = "approved"

    # Add comment if provided
    if comment:
        new_comment = QualityIssueComment(
            quality_issue_id=issue_id,
            user_id=current_user.id,
            comment=f"Issue approved: {comment}",
        )
        db.add(new_comment)

    db.commit()
    db.refresh(issue)

    # Transform affected_records from dictionary to list format if needed
    issue = transform_affected_records([issue])[0]

    return issue


@router.post("/{portfolio_id}/approve-all-quality-issues", 
             description="Approve all open quality issues for a portfolio at once", 
             response_model=Dict,
             responses={404: {"description": "Portfolio not found"},
                        401: {"description": "Not authenticated"}},)
def approve_all_quality_issues(
    portfolio_id: int,
    comment: Optional[str] = None,
    db: Session = Depends(get_tenant_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Approve all open quality issues for a portfolio at once.
    Optionally add the same comment to all issues.
    """
    # Verify portfolio exists and belongs to current user
    portfolio = (
        db.query(Portfolio)
        .filter(Portfolio.id == portfolio_id)
        .first()
    )

    if not portfolio:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found"
        )

    # Get all open quality issues
    open_issues = (
        db.query(QualityIssue)
        .filter(
            QualityIssue.portfolio_id == portfolio_id, QualityIssue.status == "open"
        )
        .all()
    )

    if not open_issues:
        return {"message": "No open quality issues to approve", "count": 0}

    # Update all issues to approved
    for issue in open_issues:
        issue.status = "approved"

        # Add comment if provided
        if comment:
            new_comment = QualityIssueComment(
                quality_issue_id=issue.id,
                user_id=current_user.id,
                comment=f"Batch approval: {comment}",
            )
            db.add(new_comment)

    db.commit()

    return {"message": "All quality issues approved", "count": len(open_issues)}


@router.post("/{portfolio_id}/recheck-quality", 
             description="Run quality checks again to find any new issues.", 
             response_model=QualityCheckSummary,
             responses={404: {"description": "Portfolio not found"},
                        401: {"description": "Not authenticated"},},)
def recheck_quality_issues(
    portfolio_id: int,
    db: Session = Depends(get_tenant_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Run quality checks again to find any new issues.
    """
    # Verify portfolio exists and belongs to current user
    portfolio = (
        db.query(Portfolio)
        .filter(Portfolio.id == portfolio_id)
        .first()
    )

    if not portfolio:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Portfolio not found"
        )

    # Run quality checks and create issues if necessary
    quality_counts = create_quality_issues_if_needed(db, portfolio_id)

    return QualityCheckSummary(
        duplicate_customer_ids=quality_counts["duplicate_customer_ids"],
        duplicate_addresses=quality_counts["duplicate_addresses"],
        duplicate_dob=quality_counts["duplicate_dob"],
        duplicate_loan_ids=quality_counts["duplicate_loan_ids"],
        unmatched_employee_ids=quality_counts["clients_without_matching_loans"],
        loan_customer_mismatches=quality_counts["loans_without_matching_clients"],
        missing_dob=quality_counts["missing_dob"],
        total_issues=quality_counts["total_issues"],
        high_severity_issues=quality_counts["high_severity_issues"],
        open_issues=quality_counts["open_issues"],
    )


@router.get("/{portfolio_id}/quality-issues/{issue_id}/download", 
            description="Download a specific quality issue as Excel", 
            status_code=status.HTTP_200_OK,
            responses={404: {"description": "Portfolio not found"},
                       401: {"description": "Not Aunthenticated"},}
                       )
async def download_quality_issue_excel(
    portfolio_id: int,
    issue_id: int,
    include_comments: bool = True,
    db: Session = Depends(get_tenant_db),
    current_user: User = Depends(get_current_active_user),
):
    # ---- Validate ownership
    portfolio = (
        db.query(Portfolio)
        .filter(
            Portfolio.id == portfolio_id,
            Portfolio.user_id == current_user.id
        )
        .first()
    )
    if not portfolio:
        raise HTTPException(404, "Portfolio not found")

    issue = (
        db.query(QualityIssue)
        .filter(
            QualityIssue.id == issue_id,
            QualityIssue.portfolio_id == portfolio_id
        )
        .first()
    )
    if not issue:
        raise HTTPException(404, "Quality issue not found")

    wb = Workbook(write_only=True)

    # ---- Issue details sheet
    ws = wb.create_sheet("Issue")

    ws.append([
        "ID", "Type", "Description",
        "Severity", "Status", "Created", "Updated"
    ])

    ws.append([
        issue.id,
        issue.issue_type,
        issue.description,
        issue.severity,
        issue.status,
        issue.created_at,
        issue.updated_at
    ])

    # ---- Affected records sheet
    if issue.affected_records and isinstance(issue.affected_records, list):
        ws_records = wb.create_sheet("Affected Records")

        headers = issue.affected_records[0].keys()
        ws_records.append(list(headers))

        for record in issue.affected_records:
            ws_records.append(excel_safe(record.values()))

    # ---- Comments sheet
    if include_comments:
        ws_comments = wb.create_sheet("Comments")

        ws_comments.append([
            "Comment ID", "User Email",
            "Comment", "Created"
        ])

        comments_query = (
            db.query(
                QualityIssueComment.id,
                User.email,
                QualityIssueComment.comment,
                QualityIssueComment.created_at
            )
            .join(User, User.id == QualityIssueComment.user_id)
            .filter(QualityIssueComment.quality_issue_id == issue_id)
        )

        for row in comments_query.yield_per(1000):
            ws_comments.append(excel_safe(row))

    # ---- Save buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"quality_issue_{issue_id}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )