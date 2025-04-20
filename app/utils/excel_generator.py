from datetime import date
from typing import Dict, Any, List, Iterable, Optional
from io import BytesIO
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
import os
from datetime import datetime
import traceback


def create_report_excel(
    portfolio_name: str,
    report_type: str,
    report_date: date,
    report_data: Dict[str, Any],
    # ** NOTE: No iterator or temp_file_path arguments here **
) -> BytesIO:
    """
    Generate an Excel report for NON-DETAILED types or from simple summary data.

    Detailed reports (ECL Detailed, Local Impairment Detailed) MUST be generated
    by their specific functions in report_generators.py which handle streaming.
    Calling this function directly for those types is an error or will result
    in an empty/error report.

    Args:
        portfolio_name: Name of the portfolio
        report_type: Type of the report
        report_date: Date of the report
        report_data: Summary data to include in the report

    Returns:
        BytesIO: Excel file as a bytes buffer
    """
    wb = load_excel_template(report_type) # Still load template based on type

    # --- MODIFICATION START ---
    # Check for detailed report types - these cannot be generated directly by this function.
    # The specific generators in report_generators.py handle them entirely.
    if report_type.lower() in ["ecl_detailed_report", "local_impairment_detailed_report"]:
        # Return an empty/error workbook because this function lacks the necessary iterator.
        print(f"Error: create_report_excel cannot generate '{report_type}'. "
              f"It must be generated via generate_ecl_detailed_report or "
              f"generate_local_impairment_details_report in report_generators.py.")
        ws = wb.active
        ws.title = "Error"
        ws['A1'] = f"Error: Report type '{report_type}' requires specific generator."
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    # --- MODIFICATION END ---

    # Handle OTHER specialized (non-streaming) report types that only need summary data
    elif report_type.lower() == "ecl_report_summarised_by_stages":
        # These functions correctly receive only the summary data
        return populate_ecl_report_summarised(wb, portfolio_name, report_date, report_data)
    elif report_type.lower() == "local_impairment_report_summarised_by_stages":
        return populate_local_impairment_report_summarised(wb, portfolio_name, report_date, report_data)
    elif report_type.lower() == "journals_report":
        # Assuming journal report only needs summary data
        return populate_journal_report(wb, portfolio_name, report_date, report_data)

    # --- Default handling for other GENERIC report types ---
    # This part processes report_data assuming it's simple keys/values,
    # lists, or dictionaries suitable for the original generic processing loop.
    print(f"Handling generic report type '{report_type}' in create_report_excel...")
    ws = wb.active
    ws.title = "Summary"

    # Define styles
    title_font = Font(name='Calibri', size=16, bold=True)
    subtitle_font = Font(name='Calibri', size=14, bold=True)
    header_font = Font(name='Calibri', size=12, bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Add title
    ws['A1'] = f"{portfolio_name} - {report_type} Report"
    ws['A1'].font = title_font; ws.merge_cells('A1:D1'); ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = f"Generated on: {report_date.strftime('%B %d, %Y')}"
    ws['A2'].font = subtitle_font; ws.merge_cells('A2:D2'); ws['A2'].alignment = Alignment(horizontal='center')

    current_row = 4
    if isinstance(report_data, dict):
        ws.cell(row=current_row, column=1, value="Summary").font = subtitle_font; current_row += 1

        # First pass: simple key-values
        for key, value in report_data.items():
            # Skip complex types or keys handled later/specifically
            if key in ['reporting_date', 'file', 'loans', 'portfolio_name', 'portfolio_id', 'description'] or \
               isinstance(value, (dict, list)) or \
               key.endswith("_iterator") or key.endswith("_path"): # Skip structural/internal keys
                continue
            display_key = " ".join(word.capitalize() for word in key.split("_"))
            c1 = ws.cell(row=current_row, column=1, value=display_key); c1.border = thin_border; c1.font = header_font
            c2 = ws.cell(row=current_row, column=2); c2.border = thin_border
            if isinstance(value, (int, float)):
                c2.value = value
                c2.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
            else:
                c2.value = str(value)
            current_row += 1
        current_row += 2

        # Second pass: handle distributions and lists (if present in report_data)
        for key, value in report_data.items():
            # Chart for distribution dictionaries
            if isinstance(value, dict) and "_distribution" in key:
                 # ... (keep chart generation logic as before) ...
                try:
                    chart_sheet = wb.create_sheet(title=key.replace('_', ' ').title()[:30]) # Max 31 chars
                    chart_sheet['A1'] = f"{key.replace('_', ' ').title()}"; chart_sheet['A1'].font = subtitle_font
                    chart_sheet['A3'] = "Category"; chart_sheet['B3'] = "Value"; chart_sheet['A3'].font = header_font; chart_sheet['B3'].font = header_font
                    row_idx = 4
                    for cat_key, cat_value in value.items():
                        chart_sheet.cell(row=row_idx, column=1, value=cat_key)
                        chart_sheet.cell(row=row_idx, column=2, value=float(cat_value) if isinstance(cat_value, (int, float, Decimal)) else cat_value) # Ensure numeric for chart
                        row_idx += 1

                    if len(value) > 0 and row_idx > 4:
                        chart_title_str = key.replace('_', ' ').title()
                        chart = None
                        data_ref = Reference(chart_sheet, min_col=2, min_row=3, max_row=row_idx-1)
                        cats_ref = Reference(chart_sheet, min_col=1, min_row=4, max_row=row_idx-1)

                        if len(value) <= 6: chart = PieChart()
                        else: chart = BarChart()

                        chart.add_data(data_ref, titles_from_data=True)
                        chart.set_categories(cats_ref)
                        chart.title = chart_title_str
                        chart_sheet.add_chart(chart, "D3")

                    ws.cell(row=current_row, column=1, value=f"{key.replace('_', ' ').title()}").font = subtitle_font
                    ws.cell(row=current_row, column=2, value=f"See '{chart_sheet.title}' sheet")
                    current_row += 2
                except Exception as chart_err:
                    print(f"Error creating chart for {key}: {chart_err}")
                    ws.cell(row=current_row, column=1, value=f"{key.replace('_', ' ').title()}").font = subtitle_font
                    ws.cell(row=current_row, column=2, value=f"Error generating chart")
                    current_row += 2


            # Table for lists of dictionaries
            elif isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                # Skip if it's the main loan list for detailed reports (handled elsewhere)
                if key == 'loans' and report_type.lower() in ["ecl_detailed_report", "local_impairment_detailed_report"]:
                    continue
                # ... (keep list generation logic as before) ...
                try:
                    list_title = key.replace('_', ' ').title()
                    list_sheet = wb.create_sheet(title=list_title[:30])
                    list_sheet['A1'] = list_title; list_sheet['A1'].font = subtitle_font
                    headers = list(value[0].keys())
                    for col_idx, header in enumerate(headers, 1):
                        display_header = " ".join(word.capitalize() for word in header.split("_"))
                        lc = list_sheet.cell(row=3, column=col_idx, value=display_header); lc.font = header_font; lc.fill = header_fill
                    for row_idx, item in enumerate(value, 4):
                        for col_idx, header in enumerate(headers, 1):
                            cell_value = item.get(header, "")
                            list_sheet.cell(row=row_idx, column=col_idx, value=cell_value) # Apply formatting if needed
                    for col_idx, _ in enumerate(headers, 1): list_sheet.column_dimensions[get_column_letter(col_idx)].auto_size = True

                    ws.cell(row=current_row, column=1, value=list_title).font = subtitle_font
                    ws.cell(row=current_row, column=2, value=f"See '{list_sheet.title}' sheet")
                    current_row += 2
                except Exception as list_err:
                     print(f"Error creating list sheet for {key}: {list_err}")
                     ws.cell(row=current_row, column=1, value=f"{key.replace('_', ' ').title()}").font = subtitle_font
                     ws.cell(row=current_row, column=2, value=f"Error generating list sheet")
                     current_row += 2

            # Nested simple dictionary
            elif isinstance(value, dict) and not "_distribution" in key:
                 # ... (keep nested dictionary logic as before) ...
                section_title = key.replace('_', ' ').title(); ws.cell(row=current_row, column=1, value=section_title).font = subtitle_font; current_row += 1
                for sub_key, sub_value in value.items():
                    display_sub_key = " ".join(word.capitalize() for word in sub_key.split("_"))
                    c1 = ws.cell(row=current_row, column=1, value=display_sub_key); c1.border = thin_border
                    c2 = ws.cell(row=current_row, column=2); c2.border = thin_border
                    if isinstance(sub_value, (int, float)):
                        c2.value = sub_value; c2.number_format = '#,##0.00' if isinstance(sub_value, float) else '#,##0'
                    else: c2.value=str(sub_value)
                    current_row += 1
                current_row += 1

    # Auto-size columns in the summary sheet
    for col in ['A', 'B', 'C', 'D']:
        try:
            ws.column_dimensions[col].auto_size = True
        except Exception as auto_size_err:
             print(f"Could not auto-size column {col}: {auto_size_err}")


    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# def populate_ecl_detailed_report(
#     wb: Workbook, 
#     portfolio_name: str, 
#     report_date: date, 
#     report_data: Dict[str, Any]
# ) -> BytesIO:
#     """
#     Populate the ECL detailed report template with data.
    
#     Args:
#         wb: Excel workbook (template)
#         portfolio_name: Name of the portfolio
#         report_date: Date of the report
#         report_data: Data for the report
        
#     Returns:
#         BytesIO: Excel file as bytes buffer
#     """
#     # Get the main worksheet (should be the first one)
#     ws = wb.active
    
#     # Define number formats
#     currency_format = '#,##0.00'
#     percentage_format = '0.00%'
    
#     # Populate header information
#     ws['B3'] = report_date
#     ws['B4'] = report_data.get('report_run_date', datetime.now().strftime("%Y-%m-%d"))
#     ws['B6'] = report_data.get('description', f"ECL Detailed Report for {portfolio_name}")
    
#     # Populate summary values
#     ws['B9'] = report_data.get('total_ead', 0)
#     ws['B9'].number_format = currency_format
    
#     ws['B10'] = report_data.get('total_lgd', 0)
#     ws['B10'].number_format = currency_format
    
#     ws['B12'] = report_data.get('total_ecl', 0)
#     ws['B12'].number_format = currency_format
    
#     # Check if there's a total loan count to display
#     if 'total_loan_count' in report_data:
#         ws['B14'] = f"Total Loans: {report_data['total_loan_count']}"
    
#     # Populate loan details starting from row 15
#     loans = report_data.get('loans', [])
#     start_row = 15
    
#     # Track temporary file path if it exists
#     temp_file_path = None
#     if hasattr(loans, 'file_path'):
#         temp_file_path = loans.file_path
    
#     # Add a counter to track progress for large datasets
#     total_loans = len(loans) if hasattr(loans, '__len__') else "unknown"
#     print(f"Processing {total_loans} loans for ECL detailed report")
    
#     # Process loans in batches to reduce memory pressure
#     batch_size = 1000
#     loan_count = 0
    
#     # Use a list comprehension to get all loans if it's an iterator
#     loan_list = list(loans) if hasattr(loans, '__iter__') else loans
    
#     for i, loan in enumerate(loan_list):
#         row = start_row + i
#         loan_count += 1
        
#         # Print progress every 1000 loans
#         if loan_count % 1000 == 0:
#             print(f"Processed {loan_count} loans...")
        
#         # Map loan data to columns
#         ws.cell(row=row, column=1, value=loan.get('loan_id', ''))
#         ws.cell(row=row, column=2, value=loan.get('employee_id', ''))
#         ws.cell(row=row, column=3, value=loan.get('employee_name', ''))
        
#         # Loan value with currency format
#         loan_value = loan.get('loan_value', 0)
#         # Handle string values (from our memory optimization)
#         if isinstance(loan_value, str):
#             try:
#                 loan_value = float(loan_value)
#             except (ValueError, TypeError):
#                 loan_value = 0
#         ws.cell(row=row, column=4, value=loan_value)
#         ws.cell(row=row, column=4).number_format = currency_format
        
#         # Outstanding loan balance with currency format
#         outstanding_balance = loan.get('outstanding_loan_balance', 0)
#         if isinstance(outstanding_balance, str):
#             try:
#                 outstanding_balance = float(outstanding_balance)
#             except (ValueError, TypeError):
#                 outstanding_balance = 0
#         ws.cell(row=row, column=5, value=outstanding_balance)
#         ws.cell(row=row, column=5).number_format = currency_format
        
#         # Accumulated arrears with currency format
#         accumulated_arrears = loan.get('accumulated_arrears', 0)
#         if isinstance(accumulated_arrears, str):
#             try:
#                 accumulated_arrears = float(accumulated_arrears)
#             except (ValueError, TypeError):
#                 accumulated_arrears = 0
#         ws.cell(row=row, column=6, value=accumulated_arrears)
#         ws.cell(row=row, column=6).number_format = currency_format
        
#         # NDIA with currency format
#         ndia = loan.get('ndia', 0)
#         if isinstance(ndia, str):
#             try:
#                 ndia = float(ndia)
#             except (ValueError, TypeError):
#                 ndia = 0
#         ws.cell(row=row, column=7, value=ndia)
#         ws.cell(row=row, column=7).number_format = currency_format
        
#         # Stage
#         ws.cell(row=row, column=8, value=loan.get('stage', ''))
        
#         # EAD with currency format
#         ead = loan.get('ead', 0)
#         if isinstance(ead, str):
#             try:
#                 ead = float(ead)
#             except (ValueError, TypeError):
#                 ead = 0
#         ws.cell(row=row, column=9, value=ead)
#         ws.cell(row=row, column=9).number_format = currency_format
        
#         # LGD as percentage
#         lgd = loan.get('lgd', 0)
#         if isinstance(lgd, str):
#             try:
#                 lgd = float(lgd) / 100  # Convert to decimal for percentage format
#             except (ValueError, TypeError):
#                 lgd = 0
#         else:
#             lgd = lgd / 100  # Convert to decimal for percentage format
#         ws.cell(row=row, column=10, value=lgd)
#         ws.cell(row=row, column=10).number_format = percentage_format
        
#         # EIR as percentage
#         eir = loan.get('eir', 0)
#         if isinstance(eir, str):
#             try:
#                 eir = float(eir)
#             except (ValueError, TypeError):
#                 eir = 0
#         ws.cell(row=row, column=11, value=eir)
#         ws.cell(row=row, column=11).number_format = percentage_format
        
#         # PD as percentage
#         pd_value = loan.get('pd', 0)
#         if isinstance(pd_value, str):
#             try:
#                 pd_value = float(pd_value)
#             except (ValueError, TypeError):
#                 pd_value = 0
#         ws.cell(row=row, column=12, value=pd_value)
#         ws.cell(row=row, column=12).number_format = percentage_format
        
#         # ECL with currency format
#         ecl = loan.get('ecl', 0)
#         if isinstance(ecl, str):
#             try:
#                 ecl = float(ecl)
#             except (ValueError, TypeError):
#                 ecl = 0
#         ws.cell(row=row, column=13, value=ecl)
#         ws.cell(row=row, column=13).number_format = currency_format
    
#     print(f"Completed processing {loan_count} loans for ECL detailed report")
    
#     # Save to BytesIO
#     buffer = BytesIO()
#     wb.save(buffer)
#     buffer.seek(0)
    
#     # Clean up temporary file if it exists
#     if temp_file_path and os.path.exists(temp_file_path):
#         try:
#             os.remove(temp_file_path)
#             print(f"Cleaned up temporary file: {temp_file_path}")
#         except Exception as e:
#             print(f"Error cleaning up temporary file: {e}")
    
#     return buffer

def populate_ecl_detailed_report(
    wb: Workbook,
    portfolio_name: str,
    report_date: date,
    report_data: Dict[str, Any], # Contains summary info like totals, run date
    loans_iterator: Iterable[Dict[str, Any]], # <-- NEW: Accepts an iterator
    temp_file_path: Optional[str] = None # <-- NEW: Path for cleanup
) -> BytesIO:
    """
    Populate the ECL detailed report template with data using an iterator.

    Args:
        wb: Excel workbook (template)
        portfolio_name: Name of the portfolio
        report_date: Date of the report
        report_data: Summary data for the report header/totals
        loans_iterator: An iterator yielding loan detail dictionaries
        temp_file_path: The path to the temporary JSON file for cleanup

    Returns:
        BytesIO: Excel file as bytes buffer
    """
    print("Starting Excel population for ECL Detailed Report using iterator...")
    ws = wb.active
    currency_format = '#,##0.00'
    percentage_format = '0.00%'

    # Populate header information from report_data
    ws['B3'] = report_date
    ws['B4'] = report_data.get('report_run_date', datetime.now().strftime("%Y-%m-%d"))
    ws['B6'] = report_data.get('description', f"ECL Detailed Report for {portfolio_name}")

    # Populate summary values from report_data
    ws['B9'] = report_data.get('total_ead', 0)
    ws['B9'].number_format = currency_format

    # Note: total_lgd in summary might represent sum(lgd * balance),
    # not just sum(lgd). Ensure report_data provides the correct summary value.
    ws['B10'] = report_data.get('total_lgd', 0)
    ws['B10'].number_format = currency_format

    ws['B12'] = report_data.get('total_ecl', 0)
    ws['B12'].number_format = currency_format


    # Populate loan details starting from row 15
    start_row = 15
    loan_count = 0
    print(f"Processing loans for Excel sheet...")

    # Iterate directly over the loan data stream
    for i, loan in enumerate(loans_iterator): # <--- CHANGE: Use the iterator directly
        row = start_row + i
        loan_count += 1

        if loan_count % 1000 == 0:
            print(f"Processed {loan_count} loans into Excel...")

        try:
            # Map loan data to columns, converting strings from JSON back to numbers
            ws.cell(row=row, column=1, value=loan.get('loan_id', ''))
            ws.cell(row=row, column=2, value=loan.get('employee_id', ''))
            ws.cell(row=row, column=3, value=loan.get('employee_name', ''))

            # Loan value
            loan_value = float(loan.get('loan_value', '0'))
            ws.cell(row=row, column=4, value=loan_value).number_format = currency_format

            # Outstanding loan balance
            outstanding_balance = float(loan.get('outstanding_loan_balance', '0'))
            ws.cell(row=row, column=5, value=outstanding_balance).number_format = currency_format

            # Accumulated arrears
            accumulated_arrears = float(loan.get('accumulated_arrears', '0'))
            ws.cell(row=row, column=6, value=accumulated_arrears).number_format = currency_format

            # NDIA
            ndia = float(loan.get('ndia', '0'))
            ws.cell(row=row, column=7, value=ndia).number_format = currency_format # Or '#,##0' if it's always integer days

            # Stage
            ws.cell(row=row, column=8, value=loan.get('stage', ''))

            # EAD
            ead = float(loan.get('ead', '0'))
            ws.cell(row=row, column=9, value=ead).number_format = currency_format

            # LGD (Stored as 'percentage value' string like '45.6')
            lgd_str = loan.get('lgd', '0')
            lgd = float(lgd_str) / 100.0 # Convert to decimal for percentage format
            ws.cell(row=row, column=10, value=lgd).number_format = percentage_format

            # EIR (Stored as 'percentage value' string)
            eir_str = loan.get('eir', '0')
            eir = float(eir_str) / 100.0
            ws.cell(row=row, column=11, value=eir).number_format = percentage_format

            # PD (Stored as 'percentage value' string)
            pd_str = loan.get('pd', '0')
            pd_value = float(pd_str) / 100.0
            ws.cell(row=row, column=12, value=pd_value).number_format = percentage_format

            # ECL
            ecl = float(loan.get('ecl', '0'))
            ws.cell(row=row, column=13, value=ecl).number_format = currency_format

        except (ValueError, TypeError) as e:
             print(f"Warning: Could not parse data for row {row}, loan_id {loan.get('loan_id', 'N/A')}. Error: {e}")
             # Optionally fill cells with error indicators or skip row
             ws.cell(row=row, column=1, value=loan.get('loan_id', 'N/A'))
             ws.cell(row=row, column=2, value="PARSE ERROR")
             # ... fill other known columns if possible ...
        except Exception as e:
            print(f"Unexpected error processing row {row}, loan_id {loan.get('loan_id', 'N/A')}: {e}")
            traceback.print_exc() # Print full traceback for debugging
            ws.cell(row=row, column=1, value=loan.get('loan_id', 'N/A'))
            ws.cell(row=row, column=2, value="PROCESSING ERROR")


    print(f"Completed processing {loan_count} loans for Excel sheet.")

    # Save to BytesIO
    buffer = BytesIO()
    try:
        print("Saving workbook to buffer...")
        wb.save(buffer)
        buffer.seek(0)
        print("Workbook saved.")
    except Exception as e:
        print(f"Error saving workbook: {e}")
        raise # Re-raise the exception

    finally:
        # Clean up temporary file INSIDE the excel generator
        # This ensures cleanup even if workbook saving fails (partially)
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
                print(f"Cleaned up temporary file: {temp_file_path}")
            except Exception as e:
                print(f"Error cleaning up temporary file '{temp_file_path}': {e}")
        else:
            if temp_file_path:
                 print(f"Temporary file path provided but file not found for cleanup: {temp_file_path}")

    return buffer


def populate_ecl_report_summarised(
    wb: Workbook, 
    portfolio_name: str, 
    report_date: date, 
    report_data: Dict[str, Any]
) -> BytesIO:
    """
    Populate the ECL summarised report template with data.
    
    Args:
        wb: Excel workbook (template)
        portfolio_name: Name of the portfolio
        report_date: Date of the report
        report_data: Data for the report
        
    Returns:
        BytesIO: Excel file as bytes buffer
    """
    # Get the main worksheet (should be the first one)
    ws = wb.active
    
    # Define number formats
    currency_format = '#,##0.00'
    percentage_format = '0.00%'
    
    # Populate header information
    ws['B3'] = report_date
    ws['B4'] = report_data.get('report_run_date', datetime.now().date())
    ws['B6'] = report_data.get('description', f"ECL Summarised Report for {portfolio_name}")
    
    # Populate stage data
    # Loan values
    ws['C18'] = report_data.get('stage_1', {}).get('loan_value', 0)
    ws['D18'] = report_data.get('stage_2', {}).get('loan_value', 0)
    ws['E18'] = report_data.get('stage_3', {}).get('loan_value', 0)
    
    # Apply currency format to loan values
    for col in ['C', 'D', 'E']:
        ws[f'{col}18'].number_format = currency_format
    
    # Outstanding loan balances
    ws['C19'] = report_data.get('stage_1', {}).get('outstanding_balance', 0)
    ws['D19'] = report_data.get('stage_2', {}).get('outstanding_balance', 0)
    ws['E19'] = report_data.get('stage_3', {}).get('outstanding_balance', 0)
    
    # Apply currency format to outstanding balances
    for col in ['C', 'D', 'E']:
        ws[f'{col}19'].number_format = currency_format
    
    # ECL amounts
    ws['C20'] = report_data.get('stage_1', {}).get('ecl', 0)
    ws['D20'] = report_data.get('stage_2', {}).get('ecl', 0)
    ws['E20'] = report_data.get('stage_3', {}).get('ecl', 0)
    
    # Apply currency format to ECL amounts
    for col in ['C', 'D', 'E']:
        ws[f'{col}20'].number_format = currency_format
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def populate_local_impairment_details_report(
    wb: Workbook,
    portfolio_name: str,
    report_date: date,
    report_data: Dict[str, Any], # Contains summary info like totals, run date
    loans_iterator: Iterable[Dict[str, Any]], # <-- NEW: Accepts an iterator
    temp_file_path: Optional[str] = None # <-- NEW: Path for cleanup
) -> BytesIO:
    """
    Populate the local impairment detailed report template using an iterator.

    Args:
        wb: Excel workbook (template)
        portfolio_name: Name of the portfolio
        report_date: Date of the report
        report_data: Summary data for the report header/totals
        loans_iterator: An iterator yielding loan detail dictionaries
        temp_file_path: The path to the temporary JSON file for cleanup

    Returns:
        BytesIO: Excel file as bytes buffer
    """
    print("Starting Excel population for Local Impairment Detailed Report using iterator...")
    ws = wb.active
    currency_format = '#,##0.00'
    percentage_format = '0.00%'

    # Populate header information from report_data
    ws['B3'] = report_date
    ws['B4'] = report_data.get('report_run_date', datetime.now().strftime("%Y-%m-%d"))
    ws['B6'] = report_data.get('description', f"Local Impairment Details Report for {portfolio_name}")

    # Populate total provision from report_data
    ws['B12'] = report_data.get('total_provision', 0)
    ws['B12'].number_format = currency_format

    # Use total_loan_count from report_data if available
    if 'total_loan_count' in report_data:
        ws['B14'] = f"Total Loans: {report_data['total_loan_count']}"
    else:
         ws['B14'] = "Total Loans: N/A"

    # Populate loan details starting from row 15
    start_row = 15
    loan_count = 0
    print(f"Processing loans for Excel sheet...")

    # Iterate directly over the loan data stream
    for i, loan in enumerate(loans_iterator): # <--- CHANGE: Use the iterator directly
        row = start_row + i
        loan_count += 1

        if loan_count % 1000 == 0:
            print(f"Processed {loan_count} loans into Excel...")

        try:
            # Map loan data to columns, converting strings from JSON back to numbers
            ws.cell(row=row, column=1, value=loan.get('loan_id', ''))
            ws.cell(row=row, column=2, value=loan.get('employee_id', ''))
            ws.cell(row=row, column=3, value=loan.get('employee_name', ''))

            # Loan value
            loan_value = float(loan.get('loan_value', '0'))
            ws.cell(row=row, column=4, value=loan_value).number_format = currency_format

            # Outstanding loan balance
            outstanding_balance = float(loan.get('outstanding_balance', '0'))
            ws.cell(row=row, column=5, value=outstanding_balance).number_format = currency_format

            # Accumulated arrears
            accumulated_arrears = float(loan.get('accumulated_arrears', '0'))
            ws.cell(row=row, column=6, value=accumulated_arrears).number_format = currency_format

            # NDIA
            ndia = float(loan.get('ndia', '0'))
            ws.cell(row=row, column=7, value=ndia).number_format = currency_format # Or '#,##0'

            # Impairment category
            ws.cell(row=row, column=8, value=loan.get('impairment_category', ''))

            # Provision rate (Stored as 'rate value' string like '0.25')
            provision_rate_str = loan.get('provision_rate', '0')
            provision_rate = float(provision_rate_str) # Already a rate, format as percentage
            ws.cell(row=row, column=9, value=provision_rate).number_format = percentage_format

            # Provision amount
            provision_amount = float(loan.get('provision_amount', '0'))
            ws.cell(row=row, column=10, value=provision_amount).number_format = currency_format

        except (ValueError, TypeError) as e:
             print(f"Warning: Could not parse data for row {row}, loan_id {loan.get('loan_id', 'N/A')}. Error: {e}")
             ws.cell(row=row, column=1, value=loan.get('loan_id', 'N/A'))
             ws.cell(row=row, column=2, value="PARSE ERROR")
        except Exception as e:
            print(f"Unexpected error processing row {row}, loan_id {loan.get('loan_id', 'N/A')}: {e}")
            traceback.print_exc()
            ws.cell(row=row, column=1, value=loan.get('loan_id', 'N/A'))
            ws.cell(row=row, column=2, value="PROCESSING ERROR")


    print(f"Completed processing {loan_count} loans for Excel sheet.")

    # Save to BytesIO
    buffer = BytesIO()
    try:
        print("Saving workbook to buffer...")
        wb.save(buffer)
        buffer.seek(0)
        print("Workbook saved.")
    except Exception as e:
        print(f"Error saving workbook: {e}")
        raise
    finally:
        # Clean up temporary file INSIDE the excel generator
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
                print(f"Cleaned up temporary file: {temp_file_path}")
            except Exception as e:
                print(f"Error cleaning up temporary file '{temp_file_path}': {e}")
        else:
             if temp_file_path:
                 print(f"Temporary file path provided but file not found for cleanup: {temp_file_path}")

    return buffer

def populate_local_impairment_report_summarised(
    wb: Workbook, 
    portfolio_name: str, 
    report_date: date, 
    report_data: Dict[str, Any]
) -> BytesIO:
    """
    Populate the local impairment summarised report template with data.
    
    Args:
        wb: Excel workbook (template)
        portfolio_name: Name of the portfolio
        report_date: Date of the report
        report_data: Data for the report
        
    Returns:
        BytesIO: Excel file as bytes buffer
    """
    # Get the main worksheet (should be the first one)
    ws = wb.active
    
    # Define number formats
    currency_format = '#,##0.00'
    percentage_format = '0.00%'
    
    # Populate header information
    ws['B3'] = report_date
    ws['B4'] = report_data.get('report_run_date', datetime.now().date())
    ws['B6'] = report_data.get('description', f"Local Impairment Summarised Report for {portfolio_name}")
    
    # Populate category data
    # Loan values
    ws['C18'] = report_data.get('current', {}).get('loan_value', 0)
    ws['D18'] = report_data.get('olem', {}).get('loan_value', 0)
    ws['E18'] = report_data.get('substandard', {}).get('loan_value', 0)
    ws['F18'] = report_data.get('doubtful', {}).get('loan_value', 0)
    ws['G18'] = report_data.get('loss', {}).get('loan_value', 0)
    
    # Apply currency format to loan values
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws[f'{col}18'].number_format = currency_format
    
    # Outstanding loan balances
    ws['C19'] = report_data.get('current', {}).get('outstanding_balance', 0)
    ws['D19'] = report_data.get('olem', {}).get('outstanding_balance', 0)
    ws['E19'] = report_data.get('substandard', {}).get('outstanding_balance', 0)
    ws['F19'] = report_data.get('doubtful', {}).get('outstanding_balance', 0)
    ws['G19'] = report_data.get('loss', {}).get('outstanding_balance', 0)
    
    # Apply currency format to outstanding balances
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws[f'{col}19'].number_format = currency_format
    
    # Provision amounts
    ws['C20'] = report_data.get('current', {}).get('provision', 0)
    ws['D20'] = report_data.get('olem', {}).get('provision', 0)
    ws['E20'] = report_data.get('substandard', {}).get('provision', 0)
    ws['F20'] = report_data.get('doubtful', {}).get('provision', 0)
    ws['G20'] = report_data.get('loss', {}).get('provision', 0)
    
    # Apply currency format to provision amounts
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws[f'{col}20'].number_format = currency_format
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def populate_journal_report(
    wb: Workbook, 
    portfolio_name: str, 
    report_date: date, 
    report_data: Dict[str, Any]
) -> BytesIO:
    """
    Populate the journal report template with data.
    
    Args:
        wb: Excel workbook (template)
        portfolio_name: Name of the portfolio (not used in this report as it handles multiple portfolios)
        report_date: Date of the report
        report_data: Data for the report
        
    Returns:
        BytesIO: Excel file as bytes buffer
    """
    # Get the main worksheet (should be the first one)
    ws = wb.active
    
    # Define styles
    bold_font = Font(name='Calibri', size=10, bold=True)
    
    # Define border for header cells
    thin_border = Border(
        bottom=Side(style='thin')
    )
    
    # Define number formats
    currency_format = '#,##0.00_);(#,##0.00)'
    
    # Populate header information
    ws['B3'] = report_date
    ws['B4'] = report_data.get('report_run_date', datetime.now().date())
    ws['B6'] = report_data.get('description', "Journal Report")
    
    # Get portfolios data
    portfolios = report_data.get('portfolios', [])
    
    # Start row for the first portfolio
    current_row = 8
    
    # Process each portfolio
    for index, portfolio in enumerate(portfolios, 1):
        # Portfolio header
        ws[f'A{current_row}'] = f"Portfolio {index}" if portfolio.get('portfolio_name') != "Summary" else "Summary"
        ws[f'A{current_row}'].font = bold_font
        current_row += 1
        
        # Column headers
        ws[f'A{current_row}'] = "Account code"
        ws[f'B{current_row}'] = "Journal description"
        ws[f'C{current_row}'] = "Amount (GHS)"
        
        # Apply bold and underline to headers
        for col in ['A', 'B', 'C']:
            ws[f'{col}{current_row}'].font = bold_font
            ws[f'{col}{current_row}'].border = thin_border
        
        current_row += 1
        
        # Journal entries
        # IFRS9 Impairment - P&L charge
        ws[f'A{current_row}'] = portfolio.get('ecl_impairment_account', '')
        ws[f'B{current_row}'] = "IFRS9 Impairment - P&l charge"
        ws[f'C{current_row}'] = portfolio.get('total_ecl', 0)
        ws[f'C{current_row}'].number_format = currency_format
        current_row += 1
        
        # IFRS9 Impairment - impact on loans
        ws[f'A{current_row}'] = portfolio.get('loan_assets', '')
        ws[f'B{current_row}'] = "IFRS9 Impairment - impact on loans"
        ws[f'C{current_row}'] = -portfolio.get('total_ecl', 0)  # Negative value
        ws[f'C{current_row}'].number_format = currency_format
        current_row += 1
        
        # Top up for BOG Impairment - P&L charge
        ws[f'A{current_row}'] = portfolio.get('ecl_impairment_account', '')
        ws[f'B{current_row}'] = "Top up for BOG Impairment - P&l charge"
        ws[f'C{current_row}'] = portfolio.get('risk_reserve', 0)
        ws[f'C{current_row}'].number_format = currency_format
        current_row += 1
        
        # Credit risk reserve
        ws[f'A{current_row}'] = portfolio.get('credit_risk_reserve', '')
        ws[f'B{current_row}'] = "Credit risk reserve"
        ws[f'C{current_row}'] = -portfolio.get('risk_reserve', 0)  # Negative value
        ws[f'C{current_row}'].number_format = currency_format
        current_row += 1
        
        # Skip a row before the next portfolio
        current_row += 1
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def load_excel_template(report_type: str) -> Workbook:
    """
    Load an Excel template for a specific report type.
    If generating very large files causes memory issues even with streaming *data*,
    consider using openpyxl's write_only mode here, but be aware of its limitations.

    Args:
        report_type: Type of the report

    Returns:
        Workbook: Excel workbook template
    """
    import os
    from openpyxl import load_workbook

    template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates", "reports")
    template_map = {
        "ecl_detailed_report": "ecl_detailed_report.xlsx",
        "ecl_report_summarised_by_stages": "ecl_report_summarised.xlsx",
        "local_impairment_detailed_report": "local_impairment_details_report.xlsx",
        "local_impairment_report_summarised_by_stages": "local_impairment_report_summarised.xlsx",
        "journals_report": "journals_report.xlsx",
    }

    template_file = template_map.get(report_type.lower())
    template_full_path = os.path.join(template_dir, template_file) if template_file else None

    # OPTIONAL: For extremely large files, uncomment write_only=True
    # This creates a workbook optimized for writing but limits reading/editing existing cells.
    # use_write_only = report_type.lower() in ["ecl_detailed_report", "local_impairment_detailed_report"]

    if template_file and os.path.exists(template_full_path):
        # return load_workbook(template_full_path, write_only=use_write_only) # If using write_only
        return load_workbook(template_full_path)
    else:
        # return Workbook(write_only=use_write_only) # If using write_only
        return Workbook()
