"""
Excel utility functions for efficient file processing.
"""
import zipfile
from io import BytesIO
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)


def count_excel_rows_fast(file_bytes: bytes) -> int:
    """
    Count data rows (excluding header) in an XLSX file efficiently.
    
    Uses a two-tier approach:
    1. Fast path: ZIP-based XML parsing (10-100x faster than pandas)
    2. Fallback: openpyxl read-only mode
    
    Args:
        file_bytes: Raw bytes of the Excel file
        
    Returns:
        Number of data rows (excluding header row)
        
    Raises:
        Exception: If file cannot be read or parsed
    """
    # ---- FAST PATH: ZIP-based XML parsing ----
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as z:
            # Excel files are ZIP archives containing XML files
            # Worksheets are in xl/worksheets/sheet*.xml
            for name in z.namelist():
                if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                    with z.open(name) as sheet:
                        # Read full content and count <row tags
                        # (line-by-line iteration fails when multiple rows are on same line)
                        content = sheet.read()
                        total_rows = content.count(b"<row")
                        # Subtract 1 for header row
                        data_rows = max(total_rows - 1, 0)
                        logger.debug(f"ZIP-based count: {total_rows} total rows, {data_rows} data rows")
                        return data_rows
    except Exception as e:
        logger.debug(f"ZIP-based counting failed: {e}, falling back to openpyxl")
    
    # ---- FALLBACK: openpyxl read-only mode ----
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        # max_row includes header, so subtract 1
        total_rows = ws.max_row if ws.max_row else 0
        data_rows = max(total_rows - 1, 0)
        wb.close()
        logger.debug(f"openpyxl count: {total_rows} total rows, {data_rows} data rows")
        return data_rows
    except Exception as e:
        logger.error(f"Failed to count Excel rows: {e}")
        raise Exception(f"Unable to read Excel file: {str(e)}")
